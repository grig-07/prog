# package/report.py

import openpyxl
from datetime import datetime

def save_report(content):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчет"

    ws['A1'] = "Отчет о расчете материалов"
    ws['A2'] = f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A4'] = content

    wb.save("report.xlsx")
